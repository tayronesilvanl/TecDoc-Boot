from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import time
from bs4 import BeautifulSoup
import openpyxl
import datetime
now = datetime.datetime.now()
date_string = now.strftime("%Y-%m-%d_%H-%M-%S")
results_sheet_name = f"Results_{date_string}"
contOk = 0

# Set up the Chrome webdriver with detach option
options = webdriver.ChromeOptions()
options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=options)

# Navigate to the TecDoc website
driver.get('https://web.tecalliance.net/tecdocsw/en/login')

# Try to find and click the accept cookies button, and handle the exception if it occurs
try:
    wait = WebDriverWait(driver, 10)
    accept_button = wait.until(EC.presence_of_element_located((By.ID, 'ppms_cm_agree-to-all')))
    accept_button.click()
except:
    pass

# Find the login and password input fields, and enter the login credentials
login_input = driver.find_element(By.ID,'userName')
login_input.send_keys('')
password_input = driver.find_element(By.ID,'password')
password_input.send_keys('')

# Click the login button
login_button = driver.find_element(By.XPATH,'//*[@id="full-width-body"]/scrollable-layout/div/div/div[1]/div[2]/div/div/div/full-width-content/div/login-form/div/div/fieldset/div/form/div[5]/div[2]/div/div/button')
login_button.click()
driver.maximize_window()
time.sleep(2)

workbook = openpyxl.load_workbook('OEdataBase.xlsx')
# Select the active worksheet
worksheet = workbook.active
#espera pagina principal carregar\
results_sheet = workbook.create_sheet(results_sheet_name)
# Write the headers to the first row of the results sheet
results_sheet.append(["Code", "Ref Number", "Brand", "Family Group", "Status"])
for row in worksheet.iter_rows(min_row=2, values_only=True):
    code = row[0]
    input_element = driver.find_element(By.ID,"part-search-input")
    Btn_Research = driver.find_element(By.XPATH,"/html/body/catalog-host/webcat-root/div[2]/webcat-header/header/div/div/div[2]/div/input-search-options-picker/div/div[4]/button")
    input_element.send_keys(code)
    Btn_Research.click()
    time.sleep(2)
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')

    try:
        wait = WebDriverWait(driver,10)
        Nodata = wait.until(EC.presence_of_element_located((By.CLASS_NAME,"text-danger")))
        print(f"NO RESULTS FOR {code}")
    except TimeoutException:
        dataok = wait.until(EC.presence_of_element_located((By.CLASS_NAME,"text-truncate")))
        RefNumber = soup.find_all("div", {"class": "text-truncate text-primary"})
        Brand = soup.find_all("span", {"class": "font-weight-bold"})
        FamilyGroup = soup.find_all("span", {"class": "generic-article"})
        Status = soup.find_all("span", {"class": "p-element"})
        print(f"salvando resultados do {code} na planilha ...")
        time.sleep(5) 
        for i in range(len(RefNumber)):
            refnumber = RefNumber[i].text
            brand = Brand[i].text
            familygroup = FamilyGroup[i].text
            results_sheet.append([code, refnumber, brand, familygroup])
            workbook.save("results.xlsx")

    driver.back()  # Move this statement outside the try block



